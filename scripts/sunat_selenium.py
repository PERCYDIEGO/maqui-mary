"""
sunat_mtc_v5.py
───────────────
Lee RUCs desde CSV y consulta:
  1. SUNAT → Razón Social, Dirección Fiscal, Estado, Condición
  2. MTC   → Código de habilitación de transporte

REQUISITOS:
    pip install selenium webdriver-manager pandas openpyxl beautifulsoup4 requests

USO:
    python sunat_mtc_v5.py
"""

import time
import requests
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuración ──────────────────────────────────────────────────
INPUT_FILE  = r"D:\proyectos_opencode\projects\Maqui-Mary\scripts\consultas_ruc.csv"
OUTPUT_FILE = r"D:\proyectos_opencode\projects\Maqui-Mary\scripts\resultado_ruc.xlsx"
COLUMNA_RUC = "RUC"
SEPARADOR   = ";"
PAUSA       = 2.0
# ───────────────────────────────────────────────────────────────────

URL_SUNAT   = "https://e-consultaruc.sunat.gob.pe/cl-ti-itmrconsruc/FrameCriterioBusquedaWeb.jsp"
URL_MTC     = "https://mtc.gob.pe/tramitesenlinea/tweb_tLinea/tw_ConsultaDGTT/Frm_rep_intra_mercancia.aspx"
URL_MTC_POST = "https://mtc.gob.pe/tramitesenlinea/tweb_tLinea/tw_ConsultaDGTT/Frm_rep_intra_mercancia_display.aspx"

MTC_SESSION = requests.Session()
MTC_SESSION.headers.update({
    "User-Agent"  : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept"      : "text/html,application/xhtml+xml",
    "Accept-Language": "es-PE,es;q=0.9",
    "Content-Type": "application/x-www-form-urlencoded",
    "Referer"     : URL_MTC,
    "Origin"      : "https://mtc.gob.pe",
})


# ══════════════════════════════════════════════════════════════════
# SUNAT — Selenium
# ══════════════════════════════════════════════════════════════════

def iniciar_chrome():
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,900")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def buscar_ruc_sunat(driver, ruc):
    try:
        driver.get(URL_SUNAT)
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.ID, "txtRuc")))
        time.sleep(0.8)
        campo = driver.find_element(By.ID, "txtRuc")
        campo.clear()
        campo.send_keys(ruc)
        time.sleep(0.3)
        driver.find_element(By.ID, "btnAceptar").click()
    except Exception:
        try:
            inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
            if inputs:
                inputs[0].clear()
                inputs[0].send_keys(ruc)
            for btn in driver.find_elements(By.CSS_SELECTOR, "input[type='button'], button"):
                txt = (btn.get_attribute("value") or btn.text or "").lower()
                if "buscar" in txt:
                    btn.click()
                    break
        except:
            pass


def esperar_sunat(driver, timeout=15):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'Domicilio Fiscal')]")
            )
        )
        return True
    except:
        return False


def extraer_sunat(driver, ruc_original):
    res = {
        "RUC_ORIGINAL"       : ruc_original,
        "RUC_SUNAT"          : "",
        "RAZON_SOCIAL"       : "",
        "TIPO_CONTRIBUYENTE" : "",
        "ESTADO_SUNAT"       : "",
        "CONDICION_SUNAT"    : "",
        "DIRECCION_FISCAL"   : "",
    }
    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")
        for item in soup.select("div.list-group-item"):
            h4s     = item.select("h4")
            valor_p = item.select_one("p")

            # Dos h4: "Número de RUC:" | "20509146668 - FABARLI S.A.C."
            if len(h4s) == 2 and not valor_p:
                label = h4s[0].get_text(strip=True).upper()
                valor = h4s[1].get_text(strip=True)
                if ("NÚMERO DE RUC" in label or "NUMERO DE RUC" in label) and " - " in valor:
                    partes = valor.split(" - ", 1)
                    res["RUC_SUNAT"]    = partes[0].strip()
                    res["RAZON_SOCIAL"] = partes[1].strip()
                continue

            # h4 + p: campos normales
            if h4s and valor_p:
                label = h4s[0].get_text(strip=True).upper()
                valor = " ".join(valor_p.get_text(strip=True).split())
                if "DOMICILIO FISCAL" in label:
                    res["DIRECCION_FISCAL"] = valor
                elif "TIPO CONTRIBUYENTE" in label:
                    res["TIPO_CONTRIBUYENTE"] = valor
                elif "ESTADO DEL CONTRIBUYENTE" in label:
                    res["ESTADO_SUNAT"] = valor
                elif "CONDICI" in label and "CONTRIBUYENTE" in label:
                    res["CONDICION_SUNAT"] = valor
    except Exception as e:
        print(f"         ⚠ Error SUNAT: {e}")
    return res


# ══════════════════════════════════════════════════════════════════
# MTC — requests (no necesita Selenium)
# ══════════════════════════════════════════════════════════════════

def obtener_viewstate_mtc():
    """Carga el formulario MTC y extrae los tokens ASP.NET."""
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        r = MTC_SESSION.get(URL_MTC, timeout=15, verify=False)
        r.encoding = "utf-8"
        soup = BeautifulSoup(r.text, "html.parser")
        tokens = {}
        for name in ["__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"]:
            el = soup.find("input", {"name": name})
            if el:
                tokens[name] = el.get("value", "")
        return tokens
    except Exception as e:
        print(f"         ⚠ Error obteniendo tokens MTC: {e}")
        return {}


def consultar_mtc(ruc):
    """Consulta MTC por RUC y devuelve lista de códigos de habilitación."""
    codigos = []
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        tokens = obtener_viewstate_mtc()
        if not tokens:
            return ["ERROR: sin tokens MTC"]

        data = {
            "__VIEWSTATE"          : tokens.get("__VIEWSTATE", ""),
            "__VIEWSTATEGENERATOR" : tokens.get("__VIEWSTATEGENERATOR", ""),
            "__EVENTVALIDATION"    : tokens.get("__EVENTVALIDATION", ""),
            "rbOpciones"           : "2",
            "txtValor"             : ruc,
            "hdopcion"             : "2",
            "hdvalore"             : ruc,
            "hdopc"                : "2",
        }

        MTC_SESSION.headers.update({"Referer": URL_MTC})
        r = MTC_SESSION.post(URL_MTC_POST, data=data, timeout=15, verify=False)
        r.encoding = "utf-8"

        soup = BeautifulSoup(r.text, "html.parser")
        tabla = soup.find("span", {"id": "lblHtml"})
        if tabla:
            filas = tabla.select("tr")
            for fila in filas[1:]:
                celdas = fila.find_all("td")
                if len(celdas) >= 2:
                    codigo = celdas[1].get_text(strip=True)
                    if codigo:
                        codigos.append(codigo)

        if not codigos:
            texto = soup.get_text()
            if "no se encontr" in texto.lower() or "sin resultado" in texto.lower():
                codigos = ["SIN REGISTRO MTC"]
            else:
                codigos = ["SIN DATOS"]

    except Exception as e:
        codigos = [f"ERROR: {str(e)[:60]}"]

    return codigos


# ══════════════════════════════════════════════════════════════════
# Excel
# ══════════════════════════════════════════════════════════════════

def exportar_excel(registros):
    df = pd.DataFrame(registros)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resultados")
        ws = writer.sheets["Resultados"]
        for cell in ws[1]:
            cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor="2F5496")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        anchos = {
            "RUC_ORIGINAL"        : 16,
            "RUC_SUNAT"           : 16,
            "RAZON_SOCIAL"        : 50,
            "TIPO_CONTRIBUYENTE"  : 30,
            "ESTADO_SUNAT"        : 22,
            "CONDICION_SUNAT"     : 18,
            "DIRECCION_FISCAL"    : 60,
            "CODIGOS_MTC"         : 25,
        }
        for i, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = anchos.get(col, 18)
        for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            fill = PatternFill("solid", fgColor="DCE6F1" if idx%2==0 else "FFFFFF")
            for cell in row:
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.freeze_panes = "A2"
    return df


# ══════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════

def main():
    print("="*57)
    print("  Consulta SUNAT + MTC — v5")
    print("="*57)

    # Leer CSV
    try:
        df = pd.read_csv(INPUT_FILE, encoding="latin1", sep=SEPARADOR, dtype=str)
        df.columns = df.columns.str.strip()
        df[COLUMNA_RUC] = df[COLUMNA_RUC].str.strip()
        rucs = df[COLUMNA_RUC].dropna().tolist()
        print(f"✅ {len(rucs)} RUCs cargados\n")
    except FileNotFoundError:
        print(f"❌ No se encontró: {INPUT_FILE}")
        input("Enter para salir..."); return
    except KeyError:
        print(f"❌ Columna '{COLUMNA_RUC}' no encontrada. Columnas: {df.columns.tolist()}")
        input("Enter para salir..."); return

    # Iniciar Chrome para SUNAT
    print("🌐 Abriendo Chrome para SUNAT...")
    driver = iniciar_chrome()
    driver.get(URL_SUNAT)

    print()
    print("┌─────────────────────────────────────────────────────┐")
    print("│  Chrome se abrió con el formulario de SUNAT.        │")
    print("│  1. Busca cualquier RUC manualmente                 │")
    print("│  2. Espera ver los resultados                       │")
    print("│  3. Vuelve aquí y presiona Enter                    │")
    print("└─────────────────────────────────────────────────────┘")
    input("\n  → Presiona Enter cuando veas resultados en Chrome: ")

    registros = []
    total = len(rucs)

    for i, ruc in enumerate(rucs):
        print(f"\n[{i+1:02d}/{total}] {ruc}")

        # ── SUNAT ──────────────────────────────────────────────────
        buscar_ruc_sunat(driver, ruc)
        encontrado = esperar_sunat(driver, timeout=12)

        if encontrado:
            data = extraer_sunat(driver, ruc)
            print(f"  SUNAT ✓  {data['RAZON_SOCIAL'][:45]}")
            print(f"           {data['DIRECCION_FISCAL'][:55]}")
        else:
            print(f"  SUNAT 🔒  No cargó — revisa Chrome y presiona Enter")
            input("  → Enter cuando veas los resultados: ")
            data = extraer_sunat(driver, ruc)

        # ── MTC ────────────────────────────────────────────────────
        codigos = consultar_mtc(ruc)
        data["CODIGOS_MTC"] = " | ".join(codigos)
        print(f"  MTC   ✓  {data['CODIGOS_MTC']}")

        registros.append(data)
        time.sleep(PAUSA)

    driver.quit()
    print("\n🔒 Chrome cerrado")

    df_result = exportar_excel(registros)

    ok_sunat = sum(1 for r in registros if r.get("RAZON_SOCIAL"))
    ok_mtc   = sum(1 for r in registros if r.get("CODIGOS_MTC") and "ERROR" not in r.get("CODIGOS_MTC","") and "SIN" not in r.get("CODIGOS_MTC",""))
    print(f"\n✅ SUNAT: {ok_sunat}/{total}  |  MTC: {ok_mtc}/{total}")
    print(f"📄 Guardado: {OUTPUT_FILE}")
    input("\nEnter para cerrar...")


if __name__ == "__main__":
    main()